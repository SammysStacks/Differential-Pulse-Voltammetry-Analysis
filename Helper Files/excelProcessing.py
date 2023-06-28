
# -------------------------------------------------------------------------- #
# ---------------------------- Imported Modules ---------------------------- #

# Basic Modules
import os
import sys
import numpy as np
import pandas as pd
from natsort import natsorted
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
from openpyxl import load_workbook, Workbook
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font

class handlingExcelFormat:   

    def __init__(self):
        # Hardcoded sheetnames for different types of excel information
        self.emptySheetName = "empty"
        self.peakInfo_SheetName = "Peak Info; File 0"
        self.signalData_Sheetname = "Signal Data; File 0"
        self.filteredData_Sheetname = "Filtered Data; File 0"
        self.baselineCurrent_Sheetname = "Baseline Current; File 0"
        self.baselineSubtractedCurrent_Sheetname = "Final Current; File 0"
        
        self.excelFolder = "Excel Files/"
        
        # Excel parameters
        self.maxAddToexcelSheet = 1048500  # Max Rows in a Worksheet
        
    def convertToXLSX(self, inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/" + self.excelFolder
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = inputExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True) # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
            else:
                # Make Excel WorkBook
                xlWorkbook = xl.Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook                
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter = excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)    
                # Save as New Excel File
                xlWorkbook.save(excelFile)
                xlWorksheets = [xlWorksheet]
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheets
    
    def splitExcelSheetsToExcelFiles(self, inputFile):
        wb = load_workbook(filename=inputFile)
        
        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value
        
            new_wb.save('{0}.xlsx'.format(sheet.title))
    
    def addExcelAesthetics(self, worksheet):
        # Initialize variables
        align = Alignment(horizontal='center',vertical='center',wrap_text=True) 
        
        # Loop through each header cell
        for headerCell in worksheet[1]:
            column_cells = worksheet[headerCell.column_letter]
            
            # Set the column width
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            worksheet.column_dimensions[headerCell.column_letter].width = max(length, worksheet.column_dimensions[headerCell.column_letter].width)
            worksheet.column_dimensions[headerCell.column_letter].bestFit = True
            # Center the Data in the Cells
            for cell in column_cells:
                cell.alignment = align
            # Set the header text color
            headerCell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return worksheet
    

class processFiles(handlingExcelFormat):
    
    def getFiles(self, dataDirectory, removeFilesContaining, analyzeFilesContaining):
        print(analyzeFilesContaining)
        # Setup parameters
        analysisFiles = []; filesAdded = set();
        
        # For each file in the directory
        for fileName in os.listdir(dataDirectory):
            fileBase = os.path.splitext(fileName)[0]
            fullPath = dataDirectory + fileName

            # Do not analyze temporary files
            if fileBase.startswith(("~",'$','.')):
                continue
            # Do not analyze files with a certain substring in the name.
            for substring in removeFilesContaining:
                if substring in fileName:
                    fileName = "."
            # Only analyze files with a certain substring in the name.
            for substring in analyzeFilesContaining:
                if substring not in fileName:
                    fileName = "."
            # Only analyze new files.
            if fileBase in filesAdded or os.path.isdir(fullPath):
                continue
            
            # Keep track of previously unseen analysis files.
            if fileName.endswith((".txt",'.csv','.xlsx')) or "." not in fileName:
                analysisFiles.append(fullPath)
                filesAdded.add(fileBase)
        
        # If not analysis files found.s
        if len(analysisFiles) == 0:
            # Stop the program.
            print("Found the Following Files:", os.listdir(dataDirectory))
            sys.exit("No TXT/CSV/XLSX Files Found in the Data Folder: " + dataDirectory)
        
        # Sort the files and return them
        analysisFiles = natsorted(analysisFiles)
        return analysisFiles
    
    class cellObject:
        def __init__(self, value):
            self.value = value

    def extractCHIData_DPV(self, chiWorksheet):
        peakCurrentList = []; peakPotentialList = []; 
        potential = []; current = []; findStart = True
        # Loop Through the Info Section and Extract the Needxed Run Info from Excel
        rowGenerator = chiWorksheet.rows
        for cell in rowGenerator:
            # Get Cell Value
            cellVal = cell[0].value
            if cellVal == None:
                continue
            # If this is Jihong's board
            if 'Command Sent' in cellVal:
                return self.extractJihongBoardData(chiWorksheet)
            elif ',' in cellVal:
                cellValues = cellVal.split(",")
                cell = []
                for cellValue in cellValues:
                    cell.append(self.cellObject(cellValue))
                cellVal = cell[0].value
            
            if findStart:
                # If Peak Found by CHI, Get Peak Potential
                if cellVal.startswith("Ep = "):
                    peakPotential = float(cellVal.split(" = ")[-1][:-1])
                    peakPotentialList.append(peakPotential)
                # If Peak Found by CHI, Get Peak Current
                elif cellVal.startswith("ip = "):
                    peakCurrent = float(cellVal.split(" = ")[-1][:-1])
                    peakCurrentList.append(peakCurrent)
                elif "Potential/V" in cellVal:
                    next(rowGenerator) # Skip Over Empty Cell After Title
                    findStart = False
            else:
                # Break out of Loop if no More Data (edge effect if someone edits excel)
                if cell[0].value == None and len(potential) != 0:
                    break
                
                # Find the Potential and Current Data points
                potential.append(float(cell[0].value))
                current.append(float(cell[1].value))

        # Convert to Numpy Array
        current = np.array(current)
        potential = np.array(potential)
        
        return potential, current, peakPotentialList, peakCurrentList
    
    def extractJihongBoardData(self, chiWorksheet):
        # Initialize program information.
        programTypes = []; scanRates = []; range_ns = []
        range_ps = []; tia_gains = []
        collectData = False
        # Intiialize holds for current and potential.
        allPeakCurrents = [[]]; allPeakPotentials = [[]];
        potential = []; allCurrent = [[]]; 
        
        # Loop through each row.
        rowGenerator = chiWorksheet.rows
        for cell in rowGenerator:
            # Get Cell Value
            cellVal = cell[0].value
            # If nothing found, ignore.
            if cellVal == None:
                continue
            
            # If a new program started.
            elif 'Command Sent' in cellVal:
                # Signal new data is coming.
                assert collectData == False
                collectData = True
                if len(allCurrent[-1]) != 0:
                    allCurrent.append([])
                    allPeakCurrents.append([])
                    allPeakPotentials.append([])
            
            # If our program is running
            elif collectData:
                # Type of program found.
                if "type:" in cellVal:
                    programType = float(cellVal.split("type:")[-1])                    
                    if programType != 1: collectData = False; continue
                    programTypes.append(programType)
                # Scan rate found.
                elif"scan rate:" in cellVal:
                    scanRates.append(float(cellVal.split("scan rate:")[-1]))
                # Range N of program found.
                elif "range_n:" in cellVal:
                    range_ns.append(float(cellVal.split("range_n:")[-1]))
                # Range P of program found.
                elif "range_p:" in cellVal:
                    range_p = float(cellVal.split("range_p:")[-1])
                    if range_p != 0: collectData = False; continue
                    range_ps.append(range_p)
                # Type of program found.
                elif "tia gain:" in cellVal:
                    tia_gains.append(float(cellVal.split("tia gain:")[-1]))
                    
                # If the program stoped
                elif "Measurement Complete:" in cellVal:
                    collectData = False
                
                # Else, we have data to analyze
                else:
                    # Subtract the two next currents.
                    lowCurrent = float(cellVal)
                    highCurrent = float(next(rowGenerator)[0].value) # Get the next value
                    currentDiff = highCurrent - lowCurrent
                    # Store the current value.
                    allCurrent[-1].append(currentDiff)
        if len(allCurrent[-1]) == 0: 
            allCurrent.pop()
            allPeakCurrents.pop()
            allPeakPotentials.pop()
        allCurrent = np.array(allCurrent)*10**-6
        
        # Get voltage information
        minPotential, maxPotential = -0.004, -0.5
        # Calculate the step size
        step_size = -(minPotential - maxPotential) / (len(allCurrent[0]) - 1)
        # Generate the interpolated values
        potential = np.arange(minPotential, maxPotential + step_size,  step_size)
        assert len(potential) == len(allCurrent[0])  
        # Make more rows
        allPotentials = np.tile(potential, (len(allCurrent), 1))
                
        return allPotentials, allCurrent, allPeakPotentials, allPeakCurrents
    
    def extractCompiledAnalysis(self, excelSheet):
        # Intiialize holds for current and potential.
        allPeakCurrents = [[]]; allPeakPotentials = [[]];
        allPotential = []; allCurrent = [[]]; 
        
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                endDataCol = len(row)
                break

        # initialize holder for variables.
        allCurrent = [[] for _ in range(endDataCol-1)]
        allPotential = [[] for _ in range(endDataCol-1)]
        allPeakCurrents = [[] for _ in range(endDataCol-1)]
        allPeakPotentials = [[] for _ in range(endDataCol-1)]
        
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow-1, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            for signalInd in range(endDataCol-1):
                allCurrent[signalInd].append(float(dataRow[signalInd+1].value))
                allPotential[signalInd].append(float(dataRow[0].value))
        allCurrent = np.array(allCurrent)*10**-6
        allPotential = np.array(allPotential)

        return allPotential, allCurrent, allPeakPotentials, allPeakCurrents
    
    def getData(self, oldFile, outputFolder, testSheetNum = 0, excelDelimiter = ","):
        """
        --------------------------------------------------------------------------
        Input Variable Definitions:
            oldFile: The Path to the Excel File Containing the Data: txt, csv, xls, xlsx
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """ 
        # Ignore temporary files
        assert not oldFile.split("/")[-1].startswith(("~", "$"))
        # Check if File Exists
        if not os.path.exists(oldFile):
            sys.exit("\nThe following Input File Does Not Exist: " + oldFile)
       
        isDocument = "." not in oldFile.split("/")[-1] # Documents have no extension.        
        # Convert TXT and CSV Files to XLSX
        if oldFile.endswith((".txt", ".csv")) or isDocument:
            # Extract Filename Information
            oldFileExtension = os.path.basename(oldFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = outputFolder + self.excelFolder
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)
        
            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, xlWorksheet = self.convertToExcel(oldFile, excelFile, excelDelimiter = excelDelimiter, overwriteXL = True, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif oldFile.endswith(".xlsx"):
            excelFile = oldFile
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum:]
        else:
            sys.exit("\nThe Following File is Neither CSV, TXT, Nor XLSX: " + oldFile)
        
        # Extract the Data
        print("\nExtracting Data from the Excel File:", excelFile)
        if xlWorksheet[0].title == self.signalData_Sheetname:
            potential, current, peakPotentialList, peakCurrentList = self.extractCompiledAnalysis(xlWorksheet[0])
        else:
            potential, current, peakPotentialList, peakCurrentList = self.extractCHIData_DPV(xlWorksheet[0])
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        return potential, current, peakPotentialList, peakCurrentList
    
    def getAllData(self, allFiles, outputFolder, testSheetNum = 0, excelDelimiter = ","):
        # Initialize holders for file data.
        allPeakPotentials = []; allPeakCurrents = []
        allPotential = []; allCurrent = []
        fileNames = []
        
        # For each file.
        for analysisFile in allFiles:
            fileName = os.path.splitext(os.path.basename(analysisFile))[0]
            
            # Extract the new data.
            potential, current, peakPotentialList, peakCurrentList = self.getData(analysisFile, outputFolder, testSheetNum, excelDelimiter)
            # Add the new incoming data.
            if isinstance(potential[0], (list, np.ndarray)):
                allCurrent.extend(current)
                allPotential.extend(potential)
                allPeakCurrents.extend(peakCurrentList)
                allPeakPotentials.extend(peakPotentialList)
                fileNames.extend([fileName + f"_{trialInd}" for trialInd in range(len(potential))])
            else:
                fileNames.append(fileName)
                allCurrent.append(current)
                allPotential.append(potential)
                allPeakCurrents.append(peakCurrentList)
                allPeakPotentials.append(peakPotentialList)
            
        print("\nFinished Compiling all the Data")
        return allPotential, allCurrent, allPeakPotentials, allPeakCurrents, fileNames
            

class saveExcelData(handlingExcelFormat):
    
    def getExcelDocument(self, excelFile, overwriteSave = False):
        # If the excel file you are saving already exists.
        if os.path.isfile(excelFile):
            # If You Want to Overwrite the Excel.
            if overwriteSave:
                print("\t\tDeleting Old Excel Workbook")
                os.remove(excelFile) 
            else:
                print("\t\tNot overwriting the file ... but your file already exists??")
            
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\t\tCreating New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            worksheet = WB.active 
            worksheet.title = self.emptySheetName
        else:
            print("\t\tExcel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            worksheet = WB.create_sheet(self.emptySheetName)
        return WB, worksheet
    
    def saveDataDPV(self, potential, current, baselineCurrent, baselineSubtractedCurrent, peakCurrents, peakPotentials, saveExcelPath):
        print("\tSaving the analysis on the file")
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Create the path to save the excel file.
        os.makedirs(os.path.dirname(saveExcelPath), exist_ok=True) # Create Output File Directory to Save Data: If None Exists
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(saveExcelPath, overwriteSave = True)
        
        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #   
        
        # Get the Header for the Data
        header = ["Potential (V)", "Recorded Current (uAmps)"]
        # Add baseline headers if availible.
        if len(baselineCurrent) != 0:
            header.extend(["Baseline Current (uAmps)", "Baseline Subtracted Current (uAmps)"])
        # Add peak information to the headers.
        header.extend(["", "Peak Potentials (V)", "Peak Currents (uAmps)"])
            
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, len(potential), self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.signalData_Sheetname
            worksheet.append(header)  # Add the header labels to this specific file.
                        
            # Loop through all data to be saved within this sheet in the excel file.
            for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(potential))):
                # Organize all the data
                row = [potential[dataInd], current[dataInd]]
                if len(baselineCurrent) != 0:
                    row.extend([baselineCurrent[dataInd], baselineSubtractedCurrent[dataInd]])
                # Add peak information if present.
                if dataInd < len(peakPotentials):
                    row.extend(["", peakPotentials[dataInd], peakCurrents[dataInd]])
                
                # Add the row to the worksheet
                worksheet.append(row)
    
            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)

        # ------------------------------------------------------------------ #
        # ---------------- Add peak information to document ---------------- # 
        
        # Start a new sheet.
        worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Get the header for the peak information
        header = ["Peak Potentials (V)", "Peak Currents (uAmps)"]
            
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, len(peakPotentials), self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.peakInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.
                        
            # Loop through all data to be saved within this sheet in the excel file.
            for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(peakPotentials))):
                # Organize all the data
                row = [peakPotentials[dataInd], peakCurrents[dataInd]]
                
                # Add the row to the worksheet
                worksheet.append(row)
    
            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
        
        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(saveExcelPath)
        WB.close()
        
    def saveAllData(self, analysisInfo, peakInfo, saveExcelPath):
        print("\nSaving a Compiled Analysis")
        # Extract the data.
        allPeakPotentials, allPeakCurrents = np.array(peakInfo, dtype=object).T
        allPotentials, allUnfilteredCurrents, allCurrents, allBaselineCurrents, allBaselineSubtractedCurrents = np.asarray(analysisInfo).transpose((1,0,2))
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Create the path to save the excel file.
        os.makedirs(os.path.dirname(saveExcelPath), exist_ok=True) # Create Output File Directory to Save Data: If None Exists
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(saveExcelPath, overwriteSave = True)
        
        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #   
        
        # Get the Header for the Data
        header = ["Potential (V)"]
        for signalInd in range(len(allCurrents)):
            header.append(f"Recorded Signal {signalInd}")
        
        potential = allPotentials[0]
        compiledSignals = np.array([allUnfilteredCurrents, allCurrents, allBaselineCurrents, allBaselineSubtractedCurrents])
        sheetNames = [self.signalData_Sheetname, self.filteredData_Sheetname, self.baselineCurrent_Sheetname, self.baselineSubtractedCurrent_Sheetname]
        # For each type of signal
        for tabInd in range(len(compiledSignals)):
            # Loop through/save all the data in batches of maxAddToexcelSheet.
            for firstIndexInFile in range(0, len(potential), self.maxAddToexcelSheet):
                # Add the information to the page
                worksheet.title = sheetNames[tabInd]
                worksheet.append(header)  # Add the header labels to this specific file.
                            
                # Loop through all data to be saved within this sheet in the excel file.
                for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(potential))):
                    # Organize all the data
                    row = [potential[dataInd]]
                    row.extend(compiledSignals[tabInd, :, dataInd])
                    
                    # Add the row to the worksheet
                    worksheet.append(row)
        
                # Finalize document
                worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
                worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
            
        # ------------------------------------------------------------------ #
        # ---------------- Add peak information to document ---------------- # 
        
        # Start a new sheet.
        worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Get the header for the peak information
        header = []
        for signalInd in range(len(allPotentials)):
            header.extend([f"Peak Potentials for {signalInd} (V)", f"Peak Currents for {signalInd} (uAmps)"])
            
        maxNumPeaks = len(max(allPeakPotentials, key = lambda peakPotentials: len(peakPotentials)))
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, maxNumPeaks, self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.peakInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.
                        
            # Loop through all data to be saved within this sheet in the excel file.
            for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, maxNumPeaks)):
                row = []
                # Organize all the data
                for signalInd in range(len(allPeakPotentials)):
                    if dataInd < len(allPeakPotentials[signalInd]):
                        row.extend([allPeakPotentials[signalInd][dataInd], allPeakCurrents[signalInd][dataInd]])
                    else:
                        row.extend([None, None])
                
                # Add the row to the worksheet
                worksheet.append(row)
    
            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
        
        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(saveExcelPath)
        WB.close()
    

    